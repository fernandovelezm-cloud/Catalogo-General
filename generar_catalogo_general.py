#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys, json, base64
from pathlib import Path
from datetime import datetime
try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet')
    import openpyxl
try:
    import requests
except ImportError:
    os.system(f'"{sys.executable}" -m pip install requests --quiet')
    import requests

THIS_DIR    = Path(__file__).parent.resolve()
VENTAS_FILE = Path(r"C:\Users\ferna\Dropbox\Fevel\Ventas\Ventas Totales\Base Ventas Totales.xlsx")
FOTOS_DIR   = THIS_DIR
FOTOS_DIR2  = Path(r"C:\Users\ferna\Dropbox\Fevel\Fotografía\Costos Ficha Técnica (complemento)")
LOGO_PATH      = Path(r"C:\Users\ferna\Dropbox\Fevel\Imagen Corporativa\flor fével\Flos Superior Blanco.png")
LOGO_FLOR_PATH = Path(r"C:\Users\ferna\Dropbox\Fevel\Imagen Corporativa\flor fével\Flor sola Blanco.png")
OUTPUT_HTML    = THIS_DIR / "Catalogo General.html"
OUTPUT_INDEX   = THIS_DIR / "index.html"
MYSOFT_FILE = THIS_DIR / "inventario_mysoft.xlsx"

COLOR_MAP = {
    'negro':'#1C1C1C','blanco':'#F0F0F0','beige':'#E8D5B0','talco':'#EDE8DE',
    'crema':'#FFF8E7','rojo':'#CC2222','rojo napa':'#B52020','azul':'#1155CC',
    'azul napa':'#1044AA','cafe':'#6B4226','cafe napa':'#5A3520','miel':'#C8860A',
    'rosado':'#FF6B9D','rosado napa':'#EE5A8A','gris':'#888888','gris napa':'#909090',
    'plomo':'#708090','plata':'#B0B0B0','dorado':'#D4AF37','bronce':'#B87333',
    'vinotinto':'#7B2336','verde':'#2D7D3A','verde napa':'#256832','naranja':'#E87722',
    'amarillo':'#F0C030','morado':'#7B2D8B','folia':'#C9A87C','natural':'#D4C5A0',
    'multi':'linear-gradient(135deg,#FF6B6B,#4ECDC4,#45B7D1,#96CEB4)',
    'multicolor':'linear-gradient(135deg,#FF6B6B,#4ECDC4,#45B7D1,#96CEB4)',
    'sin iden.':'#CCCCCC','sin identificar':'#CCCCCC',
}

def sort_tallas(lst):
    def key(t):
        try: return (0, float(t))
        except ValueError: return (1, t.lower())
    return sorted(lst, key=key)

def leer_ventas_totales():
    print(f"Leyendo {VENTAS_FILE.name}...")
    wb = openpyxl.load_workbook(str(VENTAS_FILE), read_only=True, data_only=True)
    ws = wb['Ventas']
    productos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ciudad,fecha,codigo,precio,nombre,color,talla,tipo = row[0],row[2],row[7],row[14],row[16],row[12],row[13],row[32]
        if not codigo: continue
        codigo = str(codigo).strip()
        if not codigo or codigo=='None': continue
        tipo_raw = str(tipo).strip().title() if tipo else 'Sin Clasificar'
        if codigo not in productos:
            productos[codigo]={'nombre':nombre or codigo,'precio':0,'cat':tipo_raw,'colores':set(),'tallas':set(),'fecha':None,'precios_ciudad':{}}
        p = productos[codigo]
        if fecha and precio:
            try:
                pn=int(str(precio).replace(',','').replace('.','').split('.')[0])
                if p['fecha'] is None or fecha>p['fecha']:
                    p['precio']=pn; p['fecha']=fecha
                    if nombre: p['nombre']=nombre
                    p['cat']=tipo_raw
            except: pass
        if color and str(color).strip() not in ('','None','Sin Informacion'): p['colores'].add(str(color).strip())
        if talla and str(talla).strip() not in ('','None'): p['tallas'].add(str(talla).strip())
        if ciudad and fecha and precio:
            try:
                import unicodedata
                ck_raw=str(ciudad).strip()
                ck_norm=unicodedata.normalize('NFD',ck_raw)
                ck=''.join(c for c in ck_norm if unicodedata.category(c)!='Mn').title()
                pn2=int(str(precio).replace(',','').replace('.','').split('.')[0])
                if pn2>p['precios_ciudad'].get(ck,0): p['precios_ciudad'][ck]=pn2
            except: pass
    wb.close()
    print(f"  -> {len(productos)} productos")
    return productos

def leer_mysoft():
    src = MYSOFT_FILE if MYSOFT_FILE.exists() else None
    if not src:
        dl=Path.home()/'Downloads'
        cands=list(dl.glob('*nventari*.xlsx'))+list(dl.glob('*nventari*.csv'))+list(dl.glob('*ysoft*.xlsx'))+list(dl.glob('*ysoft*.csv'))
        if not cands: print("  Mysoft no encontrado"); return None
        cands.sort(key=lambda f:f.stat().st_mtime,reverse=True); src=cands[0]
        print(f"  Usando: {src.name}")
    else:
        print(f"  Usando: {src.name}")
    inv={}
    try:
        wb=openpyxl.load_workbook(str(src),read_only=True,data_only=True); ws=wb.active
        hdr=[str(c.value).lower().strip() if c.value else '' for c in next(ws.iter_rows(min_row=1,max_row=1))]
        def ci(ns):
            for n in ns:
                if n in hdr: return hdr.index(n)
            return None
        ic=ci(['nuevo codigo','codigo']); icl=ci(['color agrupado','color']); it=ci(['talla calzado','talla','tallas'])
        for row in ws.iter_rows(min_row=2,values_only=True):
            cod=str(row[ic]).strip() if ic is not None and row[ic] else ''
            if not cod or cod=='None': continue
            clr=str(row[icl]).strip() if icl is not None and row[icl] else ''
            tal=str(row[it]).strip() if it is not None and row[it] else ''
            if cod not in inv: inv[cod]={'colores':set(),'tallas':set(),'talla_colores':{},'color_tallas':{}}
            if clr and clr.lower() not in ('','none','sin informacion'): inv[cod]['colores'].add(clr)
            if tal and tal.lower() not in ('','none'):
                inv[cod]['tallas'].add(tal)
                if clr and clr.lower() not in ('','none','sin informacion'):
                    inv[cod]['talla_colores'].setdefault(tal, set()).add(clr)
                    inv[cod]['color_tallas'].setdefault(clr, set()).add(tal)
        wb.close()
    except Exception as e: print(f"  ERROR Mysoft: {e}"); return None
    print(f"  -> {len(inv)} productos en Mysoft")
    return inv

def buscar_fotos():
    fotos={}
    # Carpeta principal
    for f in FOTOS_DIR.glob('*.jpg'):
        cod=f.stem.strip()
        if cod.isdigit() or (cod and cod[0].isdigit()): fotos[cod]=(f.name,'')
    for sub in FOTOS_DIR.iterdir():
        if sub.is_dir():
            for f in sub.glob('*.jpg'):
                cod=f.stem.strip()
                if cod not in fotos and (cod.isdigit() or (cod and cod[0].isdigit())): fotos[cod]=(f'{sub.name}/{f.name}',sub.name)
    # Carpeta complemento — copiar a subcarpeta local para que GitHub las sirva
    if FOTOS_DIR2.exists():
        import shutil
        dest_dir = FOTOS_DIR / 'complemento'
        dest_dir.mkdir(exist_ok=True)
        for f in list(FOTOS_DIR2.glob('*.jpg')) + list(FOTOS_DIR2.glob('*.jpeg')):
            cod=f.stem.strip()
            if cod.isdigit() or (cod and cod[0].isdigit()):
                dest = dest_dir / (cod + '.jpg')
                shutil.copy2(str(f), str(dest))
                if cod not in fotos:
                    fotos[cod]=(f'complemento/{cod}.jpg','complemento')
    print(f"  -> {len(fotos)} fotos")
    return fotos

def _img_b64(path, max_h=80):
    """Carga un PNG, lo reduce a max_h px de alto y lo retorna como data URI."""
    try:
        from PIL import Image
    except ImportError:
        os.system(f'"{sys.executable}" -m pip install Pillow --quiet')
        from PIL import Image
    import io
    img = Image.open(str(path)).convert('RGBA')
    w, h = img.size
    if h > max_h:
        img = img.resize((int(w * max_h / h), max_h), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format='PNG', optimize=True)
    return 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode()

def logo_b64():
    if LOGO_PATH.exists():
        print(f"  Logo: {LOGO_PATH.name}")
        return _img_b64(LOGO_PATH, max_h=80)
    print("  Logo no encontrado"); return None

def flor_b64():
    if LOGO_FLOR_PATH.exists():
        return _img_b64(LOGO_FLOR_PATH, max_h=120)
    return logo_b64()


CSS = """
* { box-sizing:border-box; margin:0; padding:0; font-family:'Segoe UI',Arial,sans-serif; }
body { background:#F5F5F5; color:#1A1A1A; }
.header { background:#111111; padding:14px 28px; display:flex; align-items:center; gap:20px; flex-wrap:wrap; position:sticky; top:0; z-index:200; box-shadow:0 2px 12px rgba(0,0,0,.35); }
.logo { height:46px; object-fit:contain; }
.logo-text { font-size:22px; font-weight:800; color:#fff; letter-spacing:2px; }
.header-right { flex:1; display:flex; gap:12px; align-items:center; flex-wrap:wrap; min-width:220px; }
#search { flex:1; min-width:160px; max-width:320px; padding:9px 16px; font-size:14px; border:1.5px solid rgba(255,255,255,.2); border-radius:20px; background:rgba(255,255,255,.08); color:#fff; outline:none; }
#search::placeholder { color:rgba(255,255,255,.4); }
#search:focus { border-color:rgba(255,255,255,.6); background:rgba(255,255,255,.12); }
#gen-date { font-size:11px; color:rgba(255,255,255,.4); white-space:nowrap; }
.tabs { display:flex; gap:4px; padding:0 24px; background:#111111; flex-wrap:wrap; border-bottom:1px solid rgba(255,255,255,.1); }
.tab { border:none; background:transparent; padding:12px 20px; font-size:15px; font-weight:600; color:rgba(255,255,255,.85); cursor:pointer; letter-spacing:.3px; transition:all .15s; border-bottom:2px solid transparent; margin-bottom:-1px; }
.tab:hover { color:#fff; }
.tab.active { color:#fff; border-bottom-color:#fff; }
.tab-count { font-size:12px; color:rgba(255,255,255,.5); border-radius:10px; padding:1px 6px; margin-left:4px; font-weight:400; }
.tab.active .tab-count { color:rgba(255,255,255,.75); }
.subcats { background:#1E1E1E; border-bottom:1px solid rgba(255,255,255,.08); padding:0 24px; }
.subcats .tab { font-size:14px; font-weight:500; padding:10px 16px; color:rgba(255,255,255,.8); }
.subcats .tab:hover { color:#fff; }
.subcats .tab.active { color:#fff; border-bottom-color:#fff; }
#grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(200px,1fr)); gap:16px; padding:20px 28px; }
.card { background:#fff; border:1px solid #EBEBEB; border-radius:8px; overflow:hidden; cursor:pointer; transition:box-shadow .18s,transform .18s; display:flex; flex-direction:column; }
.card:hover { box-shadow:0 6px 24px rgba(0,0,0,.10); transform:translateY(-2px); }
.card-img-wrap { position:relative; width:100%; padding-top:100%; background:#F8F8F8; overflow:hidden; }
.card-img-wrap img { position:absolute; top:0; left:0; width:100%; height:100%; object-fit:cover; display:block; }
.agotado-badge { position:absolute; bottom:6px; left:6px; background:rgba(0,0,0,.55); color:#fff; font-size:10px; font-weight:700; letter-spacing:.5px; padding:3px 8px; border-radius:4px; text-transform:uppercase; }
.card-body { padding:10px 12px 12px; flex:1; display:flex; flex-direction:column; gap:6px; }
.cod-badge { font-size:12px; font-weight:700; color:#555; letter-spacing:.5px; }
.precio { font-size:15px; font-weight:800; color:#1A1A1A; }
.colors-row { display:flex; flex-wrap:wrap; gap:4px; }
.color-chip { width:14px; height:14px; border-radius:50%; border:1px solid rgba(0,0,0,.12); display:inline-block; flex-shrink:0; }
.tallas-row { display:flex; flex-wrap:wrap; gap:3px; margin-top:2px; }
.talla-chip { font-size:10px; font-weight:600; color:#444; background:#F0F0F0; border-radius:4px; padding:2px 6px; }
#empty { display:none; text-align:center; padding:60px 20px; color:#AAA; font-size:15px; }
.toggle-wrap { display:flex; align-items:center; gap:6px; cursor:pointer; user-select:none; white-space:nowrap; }
.toggle-wrap input { display:none; }
.toggle-slider { position:relative; width:34px; height:18px; background:rgba(255,255,255,.25); border-radius:9px; transition:background .2s; flex-shrink:0; }
.toggle-slider::after { content:''; position:absolute; width:14px; height:14px; border-radius:50%; background:#fff; top:2px; left:2px; transition:transform .2s; box-shadow:0 1px 3px rgba(0,0,0,.2); }
.toggle-wrap input:checked + .toggle-slider { background:rgba(255,255,255,.85); }
.toggle-wrap input:checked + .toggle-slider::after { transform:translateX(16px); background:#111; }
.toggle-label { font-size:12px; color:rgba(255,255,255,.7); font-weight:500; }
select.hdr-sel { padding:7px 12px; font-size:13px; font-weight:500; border:1.5px solid rgba(255,255,255,.2); border-radius:20px; background:rgba(255,255,255,.08); color:#fff; cursor:pointer; outline:none; }
select.hdr-sel option { background:#222; color:#fff; }
#overlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,.55); z-index:500; align-items:center; justify-content:center; }
#overlay.open { display:flex; }
#modal { background:#fff; border-radius:12px; max-width:480px; width:92%; max-height:90vh; overflow-y:auto; padding:24px; position:relative; box-shadow:0 20px 60px rgba(0,0,0,.25); }
#modal-close { position:absolute; top:14px; right:16px; background:none; border:none; font-size:24px; cursor:pointer; color:#888; line-height:1; }
#modal-img { width:100%; border-radius:8px; margin-bottom:16px; display:none; }
#modal-cod { font-size:13px; color:#888; font-weight:600; margin-bottom:4px; }
#modal-nombre { font-size:16px; font-weight:700; margin-bottom:8px; }
#modal-precio { font-size:22px; font-weight:800; color:#1A1A1A; margin-bottom:12px; }
#modal-colors-label, #modal-tallas-label { font-size:11px; color:#AAA; font-weight:600; text-transform:uppercase; letter-spacing:.5px; margin-bottom:6px; }
#modal-colors { display:flex; flex-wrap:wrap; gap:8px; margin-bottom:14px; }
.modal-color-chip { display:flex; align-items:center; gap:6px; font-size:13px; }
.modal-color-dot { width:18px; height:18px; border-radius:50%; border:1px solid rgba(0,0,0,.12); flex-shrink:0; }
#modal-tallas { display:flex; flex-wrap:wrap; gap:6px; }
.modal-talla-chip { font-size:12px; font-weight:600; color:#444; background:#F0F0F0; border-radius:6px; padding:4px 10px; }
"""


MYSOFT_USER = 'fdovelez'
MYSOFT_PASS = '1379'
MYSOFT_BASE = 'https://fevel.mysoft.live'

def descargar_inventario_mysoft():
    """Inicia sesion en Mysoft y descarga el Excel de inventarios."""
    print("  Conectando a Mysoft...")
    s = requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36',
        'Referer': MYSOFT_BASE + '/login',
    })
    # Obtener pagina de login (captura cookies iniciales / CSRF si existen)
    try:
        r = s.get(MYSOFT_BASE + '/login', timeout=15)
    except Exception as e:
        print(f"  ERROR conectando a Mysoft: {e}")
        return False

    # Buscar token CSRF en el HTML
    import re
    csrf = re.search(r'name=["\']csrf_token["\'][^>]*value=["\']([^"\']+)["\']', r.text)
    if not csrf:
        csrf = re.search(r'value=["\']([^"\']+)["\'][^>]*name=["\']csrf_token["\']', r.text)
    payload = {'correo': MYSOFT_USER, 'contrasena': MYSOFT_PASS}
    if csrf:
        payload['csrf_token'] = csrf.group(1)

    # Login
    r2 = s.post(MYSOFT_BASE + '/classes/login', data=payload, timeout=15)
    # Verificar que redirigió al inventario
    r3 = s.get(MYSOFT_BASE + '/inventarios', timeout=15)
    if '/login' in r3.url:
        print("  ERROR: Login fallido. Verifica usuario/clave en el script.")
        return False
    print("  Login OK.")

    # Paso 1: pedir que genere el Excel
    r4 = s.post(MYSOFT_BASE + '/classes/excel_inventarios', data={}, timeout=60)
    if r4.status_code != 200:
        print(f"  ERROR generando Excel (HTTP {r4.status_code})")
        return False

    # Si ya es un Excel, guardarlo directo
    if r4.content[:4] in (b'PK\x03\x04', b'\xd0\xcf\x11\xe0'):
        MYSOFT_FILE.write_bytes(r4.content)
        print(f"  Inventario descargado: {len(r4.content):,} bytes")
        return True

    # Si devuelve JSON con la ruta del archivo
    import json as _json
    try:
        resp = _json.loads(r4.content.decode('utf-8-sig'))
        ruta = resp.get('fecha','')
        if not ruta:
            print(f"  ERROR: JSON sin ruta. Respuesta: {r4.text[:200]!r}")
            return False
        # Construir URL absoluta: la ruta viene como "../recursos/..." relativa a /classes/
        ruta_limpia = ruta.replace('\\','/')
        if ruta_limpia.startswith('../'):
            ruta_limpia = ruta_limpia[3:]  # quitar "../"
        url_excel = MYSOFT_BASE + '/' + ruta_limpia
        print(f"  Descargando desde: {url_excel}")
        r5 = s.get(url_excel, timeout=60)
        if r5.status_code == 200 and r5.content[:4] in (b'PK\x03\x04', b'\xd0\xcf\x11\xe0'):
            MYSOFT_FILE.write_bytes(r5.content)
            print(f"  Inventario descargado: {len(r5.content):,} bytes")
            return True
        else:
            print(f"  ERROR descargando archivo (HTTP {r5.status_code})")
            return False
    except Exception as e:
        print(f"  ERROR parseando respuesta: {e}. Respuesta: {r4.text[:200]!r}")
        return False

def main():
    print("=" * 55)
    print("  Catalogo General Fevel")
    print("=" * 55)
    print("\n[1] Descargando inventario Mysoft...")
    ok = descargar_inventario_mysoft()
    if not ok:
        if MYSOFT_FILE.exists():
            print("  Usando inventario anterior (descarga fallida).")
        else:
            print("  ADVERTENCIA: Sin inventario Mysoft. El catalogo no tendra stock.")
    print("\n[2] Leyendo ventas...")
    productos  = leer_ventas_totales()
    print("\n[4] Leyendo Mysoft...")
    inventario = leer_mysoft()
    print("\n[5] Fotos...")
    fotos      = buscar_fotos()
    print("\n[6] Logo...")
    logo_src  = logo_b64()
    flor_src  = flor_b64()

    codigos    = set(fotos.keys())
    print(f"\n  Con foto: {len(codigos)}")
    if inventario:
        print(f"  Con stock: {len(set(inventario)&codigos)} de {len(codigos)}")

    SUBCAT_CALZADO = {'51':'Sandalias Planas','52':'Valetas','54':'Tenis','55':'Sandalias Tacon','56':'Zapatillas','57':'Botines'}
    print("\n[5] Construyendo...")
    items=[]; sin_foto=0
    for i,cod in enumerate(sorted(codigos)):
        p=productos.get(cod); fe=fotos.get(cod)
        foto_rel=fe[0] if fe else None; foto_sub=fe[1] if fe else ''
        if not p and not foto_rel: continue
        colores=set(); tallas=set()
        if inventario and cod in inventario:
            colores=inventario[cod]['colores']; tallas=inventario[cod]['tallas']
            talla_colores_raw=inventario[cod].get('talla_colores',{})
        else:
            talla_colores_raw={}
        RUIDO_C=('sin iden.','sin identificar','','none','sin informacion')
        colores={c for c in colores if c.lower() not in RUIDO_C} or set()
        tallas ={t for t in tallas  if t.lower() not in ('','none','sin informacion')}
        talla_colores={t: sort_tallas([c for c in cs if c.lower() not in RUIDO_C])
                       for t,cs in talla_colores_raw.items() if t.lower() not in ('','none','sin informacion')}
        color_tallas_raw=inventario[cod].get('color_tallas',{}) if inventario and cod in inventario else {}
        color_tallas={c: sort_tallas([t for t in ts if t.lower() not in ('','none','sin informacion')])
                      for c,ts in color_tallas_raw.items() if c.lower() not in RUIDO_C}
        nombre=(p['nombre'] if p else cod) or cod
        precio=(p['precio'] if p else 0) or 0
        cat   =(p['cat']    if p else 'Accesorios')
        # Si el código empieza con prefijo de calzado, forzar categoría Calzado
        if cod[:2] in SUBCAT_CALZADO and cat.lower() in ('sin clasificar','sin categoria','accesorios','none',''):
            cat='Calzado'
        if not foto_rel: sin_foto+=1
        if cat.lower()=='calzado': subcat=SUBCAT_CALZADO.get(cod[:2],'')
        elif cat.lower() in ('bolsos','bolso'): subcat=foto_sub
        else: subcat=''
        items.append({'_idx':i,'cod':cod,'nombre':nombre,'precio':precio,'cat':cat,'subcat':subcat,
            'foto':foto_rel or '','colores':sort_tallas(list(colores)),'tallas':sort_tallas(list(tallas)),
            'talla_colores':talla_colores,'color_tallas':color_tallas,
            'mysoft_activo':inventario is not None,'precios_ciudad':p['precios_ciudad'] if p else {}})
    print(f"  -> {len(items)} productos ({sin_foto} sin foto)")
    cats=sorted(set(x['cat'] for x in items if x['cat'] and x['cat'].lower() not in ('sin clasificar','sin categoria','bonos','none','')))
    print(f"  Categorias: {cats}")

    PJ=json.dumps(items,ensure_ascii=False)
    CJ=json.dumps(cats,ensure_ascii=False)
    CMJ=json.dumps({k:v for k,v in COLOR_MAP.items()},ensure_ascii=False)
    logo_tag=f'<img class="logo" src="{logo_src}" alt="Fevel">' if logo_src else '<span class="logo-text">Fevel</span>'
    fecha=datetime.now().strftime('%d/%m/%Y %H:%M')

    JS = f"""
const PRODUCTS={PJ};
const CATEGORIES={CJ};
(function(){{
  const bar=document.getElementById('tabs-bar');
  const keys=['todos',...CATEGORIES.map(c=>c.toLowerCase())];
  const names={{'todos':'Todos',...Object.fromEntries(CATEGORIES.map(c=>[c.toLowerCase(),c]))}};
  bar.innerHTML=keys.map((k,i)=>`<button class="tab${{i===0?' active':''}}" data-cat="${{k}}">${{names[k]}} <span class="tab-count" id="cnt-${{k.replace(/\\s+/g,'-')}}">0</span></button>`).join('');
  const counts={{todos:PRODUCTS.length}};
  PRODUCTS.forEach(p=>{{const k=(p.cat||'').toLowerCase();counts[k]=(counts[k]||0)+1;}});
  Object.entries(counts).forEach(([k,v])=>{{const el=document.getElementById('cnt-'+k.replace(/\\s+/g,'-'));if(el)el.textContent=v;}});
  bar.querySelectorAll('.tab').forEach(btn=>{{
    btn.addEventListener('click',()=>{{
      bar.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));
      btn.classList.add('active'); activeCat=btn.dataset.cat;
      const sb=document.getElementById('subcats-bar');
      if(activeCat==='calzado'||activeCat==='bolsos'){{buildSubcatBar(activeCat);sb.style.display='flex';}}
      else{{sb.style.display='none';activeSubcat='';}}
      render();
    }});
  }});
}})();
let activeCat='todos',activeSubcat='',searchQ='',showSinStock=false,filterTalla='',filterColor='',filterCiudad='sin-precio';
const CALZADO_ORDER=['Sandalias Planas','Valetas','Tenis','Sandalias Tacon','Zapatillas','Botines'];
function buildSubcatBar(cat){{
  const bar=document.getElementById('subcats-bar'); activeSubcat='';
  const prods=PRODUCTS.filter(p=>(p.cat||'').toLowerCase()===cat);
  const subcats=cat==='calzado'?CALZADO_ORDER.filter(s=>prods.some(p=>p.subcat===s)):[...new Set(prods.map(p=>p.subcat).filter(Boolean))].sort();
  bar.innerHTML=['<button class="tab active" data-sub="">Todas</button>',...subcats.map(s=>`<button class="tab" data-sub="${{s}}">${{s}} <span class="tab-count">0</span></button>`)].join('');
  subcats.forEach(s=>{{const n=prods.filter(p=>p.subcat===s).length;const b=bar.querySelector(`[data-sub="${{s}}"]`);if(b){{const sp=b.querySelector('.tab-count');if(sp)sp.textContent=n;}}}});
  bar.querySelectorAll('.tab').forEach(btn=>{{btn.addEventListener('click',()=>{{bar.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));btn.classList.add('active');activeSubcat=btn.dataset.sub;render();}});}});
}}
(function(){{
  const at=[...new Set(PRODUCTS.flatMap(p=>p.tallas))];
  at.sort((a,b)=>{{const na=parseFloat(a),nb=parseFloat(b);if(!isNaN(na)&&!isNaN(nb))return na-nb;if(!isNaN(na))return -1;if(!isNaN(nb))return 1;return a.localeCompare(b);}});
  const st=document.getElementById('filter-talla');
  at.forEach(t=>{{const o=document.createElement('option');o.value=t;o.textContent=t;st.appendChild(o);}});
  const ac=[...new Set(PRODUCTS.flatMap(p=>p.colores))].sort((a,b)=>a.localeCompare(b));
  const sc=document.getElementById('filter-color');
  ac.forEach(c=>{{const o=document.createElement('option');o.value=c;o.textContent=c;sc.appendChild(o);}});
}})();
const COLOR_MAP_JS={CMJ};
function colorStyle(n){{const k=n.toLowerCase().trim();const v=COLOR_MAP_JS[k]||COLOR_MAP_JS[k.split(' ')[0]]||'#BBBBBB';return v.startsWith('linear-gradient')?`background:${{v}}`:`background:${{v}}`;}}
function fmtPrecio(p){{if(!p)return'';return'$'+parseInt(p).toLocaleString('es-CO');}}
function render(){{
  const grid=document.getElementById('grid'),empty=document.getElementById('empty'),q=searchQ.toLowerCase();
  const visible=PRODUCTS.filter(p=>{{
    if(activeCat!=='todos'&&(p.cat||'').toLowerCase()!==activeCat)return false;
    if(activeSubcat&&p.subcat!==activeSubcat)return false;
    if(q&&!p.cod.toLowerCase().includes(q)&&!p.nombre.toLowerCase().includes(q))return false;
    const ss=p.mysoft_activo&&p.colores.length===0&&p.tallas.length===0;
    if(!showSinStock&&ss)return false;
    if(filterTalla&&!p.tallas.includes(filterTalla))return false;
    if(filterColor){{
      const colsEnTalla=filterTalla&&p.talla_colores&&p.talla_colores[filterTalla]?p.talla_colores[filterTalla]:p.colores;
      if(!colsEnTalla.includes(filterColor))return false;
    }}
    if(filterCiudad&&filterCiudad!=='sin-precio'&&!(p.precios_ciudad&&p.precios_ciudad[filterCiudad]))return false;
    return true;
  }}).sort((a,b)=>{{const na=parseFloat(a.cod),nb=parseFloat(b.cod);if(!isNaN(na)&&!isNaN(nb))return nb-na;return b.cod.localeCompare(a.cod);}});
  if(!visible.length){{grid.innerHTML='';empty.style.display='block';return;}}
  empty.style.display='none';
  grid.innerHTML=visible.map(p=>{{
    const ss=p.mysoft_activo&&p.colores.length===0&&p.tallas.length===0;
    const visibleColores=filterTalla&&p.talla_colores&&p.talla_colores[filterTalla]?p.talla_colores[filterTalla]:p.colores;
    const chips=visibleColores.map(c=>`<div class="color-chip" style="${{colorStyle(c)}}" title="${{c}}"></div>`).join('');
    const visibleTallas=filterColor&&p.color_tallas&&p.color_tallas[filterColor]?p.color_tallas[filterColor]:p.tallas;
    const talls=visibleTallas.map(t=>`<div class="talla-chip">${{t}}</div>`).join('');
    const img=p.foto?`<img src="${{p.foto}}" alt="${{p.cod}}" loading="lazy">`:'';
    const pr=filterCiudad==='sin-precio'?null:filterCiudad?(p.precios_ciudad&&p.precios_ciudad[filterCiudad]):p.precio;
    return `<div class="card${{ss?' sin-stock':''}}" onclick="openModal(${{p._idx}})">
      <div class="card-img-wrap">${{img}}${{ss?'<div class="agotado-badge">Sin stock</div>':''}}</div>
      <div class="card-body">
        <div class="cod-badge">${{p.cod}}</div>
        ${{pr?`<div class="precio">${{fmtPrecio(pr)}}</div>`:''}}
        ${{ss?'':`<div class="colors-row">${{chips}}</div>`}}
        ${{p.tallas.length?`<div class="tallas-row">${{talls}}</div>`:''}}
      </div></div>`;
  }}).join('');
}}
document.getElementById('search').addEventListener('input',e=>{{searchQ=e.target.value.trim();render();}});
document.getElementById('toggle-sin-stock').addEventListener('change',e=>{{showSinStock=e.target.checked;render();}});
document.getElementById('filter-talla').addEventListener('change',e=>{{filterTalla=e.target.value;render();}});
document.getElementById('filter-color').addEventListener('change',e=>{{filterColor=e.target.value;render();}});
document.getElementById('filter-ciudad').addEventListener('change',e=>{{filterCiudad=e.target.value;render();}});
function openModal(idx){{
  const p=PRODUCTS[idx]; if(!p)return;
  document.getElementById('modal-img').src=p.foto||'';
  document.getElementById('modal-img').style.display=p.foto?'block':'none';
  document.getElementById('modal-cod').textContent='Codigo: '+p.cod;
  document.getElementById('modal-nombre').textContent=p.nombre;
  const pr=filterCiudad==='sin-precio'?null:filterCiudad?(p.precios_ciudad&&p.precios_ciudad[filterCiudad]):p.precio;
  document.getElementById('modal-precio').textContent=pr?fmtPrecio(pr):'';
  document.getElementById('modal-colors').innerHTML=p.colores.map(c=>`<div class="modal-color-chip"><div class="modal-color-dot" style="${{colorStyle(c)}}"></div><span>${{c}}</span></div>`).join('');
  document.getElementById('modal-tallas').innerHTML=p.tallas.map(t=>`<div class="modal-talla-chip">${{t}}</div>`).join('');
  document.getElementById('modal-colors-label').style.display=p.colores.length?'':'none';
  document.getElementById('modal-tallas-label').style.display=p.tallas.length?'':'none';
  document.getElementById('overlay').classList.add('open');
}}
document.getElementById('modal-close').addEventListener('click',()=>document.getElementById('overlay').classList.remove('open'));
document.getElementById('overlay').addEventListener('click',e=>{{if(e.target===document.getElementById('overlay'))document.getElementById('overlay').classList.remove('open');}});
render();
"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Catalogo General Fevel</title>
<style>{CSS}</style>
</head>
<body>

<!-- PANTALLA DE CLAVE -->
<div id="auth-overlay" style="position:fixed;inset:0;background:#111;display:flex;flex-direction:column;align-items:center;justify-content:center;z-index:9999;">
  <img src="{flor_src}" alt="Fevel" style="height:90px;margin-bottom:28px;">
  <div style="background:#1e1e1e;border-radius:12px;padding:36px 40px;display:flex;flex-direction:column;align-items:center;gap:16px;min-width:300px;">
    <p style="color:rgba(255,255,255,.7);font-size:14px;margin:0;font-family:sans-serif;">Ingresa la clave para ver el catálogo</p>
    <input id="auth-input" type="password" placeholder="Clave de acceso"
      style="width:100%;padding:10px 14px;border-radius:8px;border:1px solid rgba(255,255,255,.2);background:#111;color:#fff;font-size:15px;outline:none;box-sizing:border-box;"
      onkeydown="if(event.key==='Enter')checkAuth()">
    <div id="auth-error" style="color:#ff6b6b;font-size:13px;display:none;font-family:sans-serif;">Clave incorrecta</div>
    <button onclick="checkAuth()"
      style="width:100%;padding:11px;border-radius:8px;border:none;background:#fff;color:#111;font-size:15px;font-weight:700;cursor:pointer;font-family:sans-serif;">
      Entrar
    </button>
  </div>
</div>
<script>
(function(){{
  const KEY='fevel1379';
  const enIframe=window.self!==window.top;
  if(enIframe||sessionStorage.getItem('fevel_auth')===KEY){{
    document.getElementById('auth-overlay').style.display='none';
  }}
  window.checkAuth=function(){{
    const v=document.getElementById('auth-input').value;
    if(v===KEY){{sessionStorage.setItem('fevel_auth',KEY);document.getElementById('auth-overlay').style.display='none';}}
    else{{document.getElementById('auth-error').style.display='block';document.getElementById('auth-input').value='';}}
  }};
}})();
</script>

<div class="header">
  {logo_tag}
  <div class="header-right">
    <input type="text" id="search" placeholder="Buscar por codigo...">
    <label class="toggle-wrap">
      <input type="checkbox" id="toggle-sin-stock">
      <span class="toggle-slider"></span>
      <span class="toggle-label">Sin stock</span>
    </label>
    <select id="filter-talla" class="hdr-sel"><option value="">Todas las tallas</option></select>
    <select id="filter-color" class="hdr-sel"><option value="">Todos los colores</option></select>
    <select id="filter-ciudad" class="hdr-sel">
      <option value="">Precio (todas)</option>
      <option value="Cartago">Precio Cartago</option>
      <option value="Bogota">Precio Bogota</option>
      <option value="sin-precio" selected>Sin precio</option>
    </select>
    <span id="gen-date">Actualizado: {fecha}</span>
  </div>
</div>
<div class="tabs" id="tabs-bar"></div>
<div class="tabs subcats" id="subcats-bar" style="display:none;"></div>
<div id="grid"></div>
<div id="empty">No se encontraron productos.</div>
<div id="overlay">
  <div id="modal">
    <button id="modal-close">&times;</button>
    <img id="modal-img" src="" alt="">
    <div id="modal-body">
      <div id="modal-cod"></div>
      <div id="modal-nombre"></div>
      <div id="modal-precio"></div>
      <div id="modal-colors-label">Colores disponibles</div>
      <div id="modal-colors"></div>
      <div id="modal-tallas-label">Tallas</div>
      <div id="modal-tallas"></div>
    </div>
  </div>
</div>
<script>{JS}</script>
</body>
</html>"""

    OUTPUT_HTML.write_text(html, encoding='utf-8')
    # index.html para GitHub Pages (redirige al catalogo)
    OUTPUT_INDEX.write_text(
        '<!DOCTYPE html><html><head><meta charset="utf-8">'
        '<meta http-equiv="refresh" content="0; url=Catalogo General.html">'
        '<script>window.location.href="Catalogo General.html";</script>'
        '</head><body></body></html>', encoding='utf-8')
    print(f"\nCatalogo generado: {OUTPUT_HTML}")
    print(f"  {len(items)} productos | {len(cats)} categorias")

if __name__ == '__main__':
    main()
